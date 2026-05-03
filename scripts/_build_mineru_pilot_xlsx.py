"""Build data/mineru_pilot_review.xlsx for the 5-case MinerU pilot."""
import json, sys, io, pathlib
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from api.neo4j_client import run_read_query

PILOT_AIDS = ['719102', '663028', '718505', '726298', '738706']

rows = run_read_query("""
  UNWIND $aids AS aid
  MATCH (a:AuctionProperty {auction_id: aid})-[:HAS_DOCUMENT]->(d:Document)
  RETURN a.auction_id          AS auction_id,
         d.filename            AS filename,
         d.file_path           AS file_path,
         a.description         AS current_desc,
         a.description_scraped AS scraped,
         a.description_source  AS current_source
""", {"aids": PILOT_AIDS}, max_rows=20)

def safe_cache_name(file_path: str) -> str:
    return file_path.replace('/', '_').replace('\\', '_').replace(':', '_')

v2_cache = pathlib.Path('pipeline/cache/notice_descriptions')
v3_cache = pathlib.Path('pipeline/cache/notice_descriptions_v3')
md_cache = pathlib.Path('pipeline/cache/mineru_markdown')

def read_json_field(p: pathlib.Path, key: str) -> str:
    if not p.exists(): return ''
    try:
        return (json.loads(p.read_text(encoding='utf-8')) or {}).get(key) or ''
    except Exception:
        return ''

by_aid = {r['auction_id']: r for r in rows}
records = []
for aid in PILOT_AIDS:
    r = by_aid.get(aid)
    if not r:
        continue
    fp = r['file_path']
    safe = safe_cache_name(fp)
    v2 = read_json_field(v2_cache / f"{safe}.json", 'property_description_full')
    v3 = read_json_field(v3_cache / f"{safe}.json", 'property_description_full')
    md_path = md_cache / f"{safe}.md"
    md = md_path.read_text(encoding='utf-8') if md_path.exists() else ''
    records.append({
        'auction_id': aid,
        'url': f'https://www.eauctionsindia.com/properties/{aid}',
        'filename': r['filename'],
        'current_source': r['current_source'] or '',
        'scraped': r['scraped'] or '',
        'v2_gemini_only': v2,
        'v3_mineru_llm': v3,
        'mineru_markdown': md,
        'len_scraped': len((r['scraped'] or '').strip()),
        'len_v2': len(v2.strip()),
        'len_v3': len(v3.strip()),
        'len_md': len(md),
    })

wb = Workbook(); ws = wb.active; ws.title = 'pilot_compare'
headers = ['auction_id', 'url', 'filename', 'current_source',
           'len_scraped', 'len_v2', 'len_v3', 'len_md',
           'scraped', 'v2_gemini_only', 'v3_mineru_llm', 'mineru_markdown']
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
for i, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = header_font; c.fill = header_fill
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for ri, e in enumerate(records, start=2):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=ri, column=ci, value=e[h])
        if h in ('scraped', 'v2_gemini_only', 'v3_mineru_llm', 'mineru_markdown'):
            c.alignment = Alignment(wrap_text=True, vertical='top')

widths = {'auction_id':10, 'url':45, 'filename':40, 'current_source':16,
          'len_scraped':10, 'len_v2':8, 'len_v3':8, 'len_md':8,
          'scraped':70, 'v2_gemini_only':70, 'v3_mineru_llm':70, 'mineru_markdown':80}
for i, h in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(i)].width = widths[h]
ws.freeze_panes = 'E2'
ws.row_dimensions[1].height = 32
for ri in range(2, len(records) + 2):
    ws.row_dimensions[ri].height = 260

target = pathlib.Path('data/mineru_pilot_review.xlsx')
fallback = pathlib.Path('data/mineru_pilot_review.v2.xlsx')
try:
    wb.save(target)
    print(f"Saved: {target}")
except PermissionError:
    wb.save(fallback)
    print(f"Original locked. Saved: {fallback}")

print(f"\nRecords: {len(records)}")
print(f"  {'aid':<8}  {'scraped':>8}  {'v2':>5}  {'v3':>5}  {'md':>6}")
for r in records:
    print(f"  {r['auction_id']:<8}  {r['len_scraped']:>8}  {r['len_v2']:>5}  {r['len_v3']:>5}  {r['len_md']:>6}")
